"""Tests for XlsxV2Adapter (v0.8.0 parser v2, spec sec.4.3).

Focus on the v2 additions over the legacy XLSX adapter:
  1. Merged cells -> HTML ``<table>`` with colspan/rowspan
  2. number_format-aware values (currency, percent, date)
  3. Empty cells rendered as em-dash
  4. Sheets emitted as HTML tables (not markdown pipe tables) so spans
     can survive the round trip
"""
from __future__ import annotations

import datetime as dt
from pathlib import Path

import pytest

from kb_extract.adapters.xlsx_v2 import XlsxV2Adapter
from kb_extract.hardness import assert_invariants

pytestmark = pytest.mark.disable_socket


def _make_basic_xlsx(path: Path) -> Path:
    import openpyxl
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Sheet1"
    ws.append(["h1", "h2"])
    ws.append(["a", "b"])
    wb.save(str(path))
    return path


def _make_merged_xlsx(path: Path) -> Path:
    import openpyxl
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Data"
    ws["A1"] = "merged-header"
    ws["A2"] = "left"
    ws["B2"] = "right"
    ws.merge_cells("A1:B1")
    wb.save(str(path))
    return path


def _make_rowspan_xlsx(path: Path) -> Path:
    import openpyxl
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Data"
    ws["A1"] = "stacked"
    ws["B1"] = "x"
    ws["B2"] = "y"
    ws.merge_cells("A1:A2")
    wb.save(str(path))
    return path


def _make_formatted_xlsx(path: Path) -> Path:
    import openpyxl
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Fmt"
    ws.append(["pct", "money", "when"])
    ws["A2"] = 0.125
    ws["A2"].number_format = "0.0%"
    ws["B2"] = 1234.5
    ws["B2"].number_format = '"$"#,##0.00'
    ws["C2"] = dt.datetime(2026, 6, 13, 9, 30)
    ws["C2"].number_format = "yyyy-mm-dd"
    wb.save(str(path))
    return path


def _make_sparse_xlsx(path: Path) -> Path:
    import openpyxl
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Sparse"
    ws["A1"] = "x"
    ws["C1"] = "z"  # B1 empty
    ws["A2"] = "a"
    wb.save(str(path))
    return path


# --- baseline parity ---------------------------------------------------------


def test_v2_basic_extracts_sheet(tmp_path: Path) -> None:
    src = _make_basic_xlsx(tmp_path / "s.xlsx")
    out = tmp_path / "out"
    (out / "assets").mkdir(parents=True)
    res = XlsxV2Adapter().extract(src, out)
    assert "# Sheet1" in res.markdown
    assert "<td>h1</td>" in res.markdown
    assert "<td>b</td>" in res.markdown


def test_v2_metadata_and_extensions(tmp_path: Path) -> None:
    src = _make_basic_xlsx(tmp_path / "s.xlsx")
    out = tmp_path / "out"
    (out / "assets").mkdir(parents=True)
    res = XlsxV2Adapter().extract(src, out)
    assert res.meta.adapter_name == "xlsx_v2"
    assert XlsxV2Adapter.extensions == (".xlsx",)


def test_v2_hardness_invariants(tmp_path: Path) -> None:
    src = _make_basic_xlsx(tmp_path / "s.xlsx")
    out = tmp_path / "out"
    (out / "assets").mkdir(parents=True)
    res = XlsxV2Adapter().extract(src, out)
    assert_invariants(res, src, out, total_pages=res.index.page_end)


def test_v2_deterministic(tmp_path: Path) -> None:
    src = _make_basic_xlsx(tmp_path / "s.xlsx")
    out1 = tmp_path / "out1"
    (out1 / "assets").mkdir(parents=True)
    out2 = tmp_path / "out2"
    (out2 / "assets").mkdir(parents=True)
    r1 = XlsxV2Adapter().extract(src, out1)
    r2 = XlsxV2Adapter().extract(src, out2)
    assert r1.markdown == r2.markdown


# --- merged cells ------------------------------------------------------------


def test_v2_horizontal_merge_emits_colspan(tmp_path: Path) -> None:
    src = _make_merged_xlsx(tmp_path / "m.xlsx")
    out = tmp_path / "out"
    (out / "assets").mkdir(parents=True)
    res = XlsxV2Adapter().extract(src, out)
    assert 'colspan="2"' in res.markdown
    assert "merged-header" in res.markdown
    assert "<td>left</td>" in res.markdown
    assert "<td>right</td>" in res.markdown


def test_v2_vertical_merge_emits_rowspan(tmp_path: Path) -> None:
    src = _make_rowspan_xlsx(tmp_path / "r.xlsx")
    out = tmp_path / "out"
    (out / "assets").mkdir(parents=True)
    res = XlsxV2Adapter().extract(src, out)
    assert 'rowspan="2"' in res.markdown
    assert "stacked" in res.markdown


# --- number formats ----------------------------------------------------------


def test_v2_percent_formatted(tmp_path: Path) -> None:
    src = _make_formatted_xlsx(tmp_path / "f.xlsx")
    out = tmp_path / "out"
    (out / "assets").mkdir(parents=True)
    res = XlsxV2Adapter().extract(src, out)
    # 0.125 with '0.0%' -> '12.5%'
    assert "12.5%" in res.markdown


def test_v2_currency_formatted(tmp_path: Path) -> None:
    src = _make_formatted_xlsx(tmp_path / "f.xlsx")
    out = tmp_path / "out"
    (out / "assets").mkdir(parents=True)
    res = XlsxV2Adapter().extract(src, out)
    assert "$1,234.50" in res.markdown


def test_v2_date_formatted(tmp_path: Path) -> None:
    src = _make_formatted_xlsx(tmp_path / "f.xlsx")
    out = tmp_path / "out"
    (out / "assets").mkdir(parents=True)
    res = XlsxV2Adapter().extract(src, out)
    assert "2026-06-13" in res.markdown


# --- empty cells -------------------------------------------------------------


def test_v2_empty_cells_render_as_em_dash(tmp_path: Path) -> None:
    src = _make_sparse_xlsx(tmp_path / "p.xlsx")
    out = tmp_path / "out"
    (out / "assets").mkdir(parents=True)
    res = XlsxV2Adapter().extract(src, out)
    # B1 was never assigned -> em-dash (U+2014) in the rendered grid
    assert "\u2014" in res.markdown
