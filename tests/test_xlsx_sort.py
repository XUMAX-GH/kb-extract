"""SP-2 / v0.2.0: XLSX deterministic natural sort by numeric prefix."""

from __future__ import annotations

import openpyxl
import pytest

from kb_extract.adapters.xlsx import XlsxAdapter, _natural_key


def test_natural_key_orders_numeric_prefixes_correctly():
    sheets = ["10_Final", "01_Intro", "02_Data", "Misc", "Aux"]
    assert sorted(sheets, key=_natural_key) == [
        "01_Intro", "02_Data", "10_Final",
        # No-numeric-prefix sheets sort after, alpha:
        "Aux", "Misc",
    ]


def test_natural_key_no_leading_digit_sorts_after_leading_digit():
    sheets = ["Z_Last", "1_First", "Alpha"]
    assert sorted(sheets, key=_natural_key) == ["1_First", "Alpha", "Z_Last"]


def test_natural_key_pure_alpha_sorts_alpha():
    sheets = ["banana", "apple", "cherry"]
    assert sorted(sheets, key=_natural_key) == ["apple", "banana", "cherry"]


@pytest.mark.disable_socket
def test_xlsx_adapter_sorts_sheets_by_numeric_prefix(tmp_path):
    wb = openpyxl.Workbook()
    # Workbook starts with one default sheet "Sheet" — rename it and add more.
    wb.active.title = "10_Z"
    wb["10_Z"]["A1"] = "z1"
    wb.create_sheet("01_A")
    wb["01_A"]["A1"] = "a1"
    wb.create_sheet("02_B")
    wb["02_B"]["A1"] = "b1"
    src = tmp_path / "deck.xlsx"
    wb.save(str(src))

    out = tmp_path / "out.tmp"
    out.mkdir()
    result = XlsxAdapter().extract(src, out)

    # Expect order 01_A, 02_B, 10_Z in the index children.
    titles = [c.title for c in result.index.children]
    assert titles == ["01_A", "02_B", "10_Z"], (
        f"Expected natural-prefix sort, got {titles}"
    )
    # page_start should increment with the sorted order (1, 2, 3)
    page_starts = [c.page_start for c in result.index.children]
    assert page_starts == [1, 2, 3]
    # And meta confidence is high (deterministic, native sheet structure)
    assert result.meta.outline_confidence == "high"
