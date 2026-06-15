"""XLSX v2 adapter — MinerU-inspired (v0.8.0 parser v2, spec sec.4.3).

Improvements over the legacy XlsxAdapter:
  - Merged cells: ``ws.merged_cells.ranges`` is used to build a
    phantom-free :class:`._table_utils.CellInfo` grid with colspan and
    rowspan, then rendered with :func:`._table_utils.cells_to_html`
    so HTML ``<table>`` markup carries the spans.
  - Number formats: ``cell.number_format`` is honoured for percent,
    currency, and date-style formats via :func:`_format_value`. Plain
    numbers/strings fall back to ``str(value)``.
  - Empty cells render as U+2014 (em-dash).

Cannot use ``read_only=True`` because that mode does not populate
``merged_cells.ranges`` reliably; we accept the memory cost for the
sake of correctness.
"""
from __future__ import annotations

import datetime as _dt
import hashlib
import re
from pathlib import Path

import openpyxl
from openpyxl import __version__ as openpyxl_version

from ..contracts import ExtractionResult, SectionNode, TableRef
from ._common import make_meta
from ._table_utils import CellInfo, cells_to_html
from .xlsx import _natural_key  # reuse stable sort

_EM_DASH = "\u2014"


def _format_value(value, number_format: str | None) -> str:
    """Render an openpyxl cell value as a string, honouring number_format.

    Heuristic-only (we deliberately don't pull in the heavy
    ``openpyxl.utils.cell.cell_format`` machinery).  Behaviour:
      * None -> em-dash
      * datetime/date/time -> ISO-ish, format depends on whether the
        number_format hints at date or time only.
      * Numeric with ``%`` in format -> ``"<value*100><decimals>%"``
      * Numeric with ``$`` in format -> ``"$X,XXX.XX"`` (decimals from
        format, default 2)
      * Otherwise ``str(value)``.
    """
    if value is None or value == "":
        return _EM_DASH

    fmt = (number_format or "").strip()

    if isinstance(value, _dt.datetime):
        if fmt and "h" in fmt.lower() and "y" not in fmt.lower():
            return value.strftime("%H:%M:%S")
        return value.strftime("%Y-%m-%d %H:%M:%S") if (
            value.hour or value.minute or value.second
        ) else value.strftime("%Y-%m-%d")
    if isinstance(value, _dt.date):
        return value.strftime("%Y-%m-%d")
    if isinstance(value, _dt.time):
        return value.strftime("%H:%M:%S")

    if isinstance(value, (int, float)) and fmt:
        # Percent formats like '0%', '0.0%', '0.00%'
        if "%" in fmt:
            decimals = 0
            m = re.search(r"0\.(0+)%", fmt)
            if m:
                decimals = len(m.group(1))
            return f"{value * 100:.{decimals}f}%"
        # Currency formats like '"$"#,##0.00' or '$#,##0'
        if "$" in fmt:
            decimals = 2 if "." in fmt else 0
            return f"${value:,.{decimals}f}"

    return str(value)


def _build_grid(ws) -> list[list[CellInfo]]:
    """Build a phantom-free grid honouring merged_cells.ranges.

    For each merged range, only the top-left cell emits a CellInfo with
    the appropriate colspan/rowspan; all other coordinates inside the
    range are dropped so :func:`cells_to_html` lays them out correctly.
    """
    # Map (row, col) -> origin (row, col) for cells inside a merge.
    origin: dict[tuple[int, int], tuple[int, int]] = {}
    span_of_origin: dict[tuple[int, int], tuple[int, int]] = {}
    for mr in ws.merged_cells.ranges:
        r0, r1 = mr.min_row, mr.max_row
        c0, c1 = mr.min_col, mr.max_col
        span_of_origin[(r0, c0)] = (r1 - r0 + 1, c1 - c0 + 1)
        for r in range(r0, r1 + 1):
            for c in range(c0, c1 + 1):
                if (r, c) != (r0, c0):
                    origin[(r, c)] = (r0, c0)

    max_row = ws.max_row or 0
    max_col = ws.max_column or 0
    grid: list[list[CellInfo]] = []
    for r in range(1, max_row + 1):
        row_out: list[CellInfo] = []
        any_nonempty = False
        for c in range(1, max_col + 1):
            if (r, c) in origin:
                continue  # covered by a merge
            cell = ws.cell(row=r, column=c)
            text = _format_value(cell.value, cell.number_format)
            if text != _EM_DASH:
                any_nonempty = True
            rowspan, colspan = span_of_origin.get((r, c), (1, 1))
            row_out.append(CellInfo(text=text, colspan=colspan, rowspan=rowspan))
        # Keep blank rows only if they sit between non-blank ones; we
        # trim trailing blanks below.
        grid.append(row_out)
        if any_nonempty:
            grid[-1] = row_out
    # Trim trailing rows that are entirely em-dash (true blanks)
    while grid and all(c.text == _EM_DASH for c in grid[-1]):
        grid.pop()
    return grid


class XlsxV2Adapter:
    name = "xlsx_v2"
    version = "0.8.0"
    extensions = (".xlsx",)

    def extract(self, src: Path, out_dir_tmp: Path) -> ExtractionResult:
        wb = openpyxl.load_workbook(str(src), data_only=True)
        sha = hashlib.sha256(src.read_bytes()).hexdigest()
        md_lines: list[str] = [
            f"<!-- generated by kb-extract; do not edit; source_sha256: {sha} -->",
            "",
        ]
        children: list[SectionNode] = []
        tables: list[TableRef] = []
        table_count = 0

        ordered_sheet_names = sorted(wb.sheetnames, key=_natural_key)
        for sheet_index, sheet_name in enumerate(ordered_sheet_names, start=1):
            ws = wb[sheet_name]
            sheet_anchor = f"sec-{sheet_index:04d}"
            md_lines.append(f'<a id="{sheet_anchor}"></a>')
            md_lines.append(f"# {sheet_name}")
            md_lines.append("")

            grid = _build_grid(ws)
            if grid:
                table_count += 1
                t_anchor = f"tbl-{table_count:04d}"
                md_lines.append(f'<a id="{t_anchor}"></a>')
                md_lines.append(cells_to_html(grid))
                md_lines.append("")
                tables.append(TableRef(
                    anchor=t_anchor, page=sheet_index,
                    rows_json=tuple(tuple(c.text for c in r) for r in grid),
                    rendered_asset=None,
                ))

            children.append(SectionNode(
                node_id=f"{sheet_index:04d}",
                title=sheet_name, level=1,
                page_start=sheet_index, page_end=sheet_index,
                anchor=sheet_anchor, language="und",
            ))
        wb.close()

        total_pages = max(len(ordered_sheet_names), 1)
        root = SectionNode(
            node_id="0000", title=src.stem, level=0,
            page_start=1, page_end=total_pages,
            anchor="", language="und", children=tuple(children),
        )
        markdown = "\n".join(md_lines) + "\n"
        meta = make_meta(
            src=src,
            adapter_name=self.name,
            adapter_version=self.version,
            tool_versions={"openpyxl": openpyxl_version},
            outline_source="heading_style",
        )
        return ExtractionResult(
            markdown=markdown, index=root,
            tables=tuple(tables), assets=(), meta=meta,
        )
